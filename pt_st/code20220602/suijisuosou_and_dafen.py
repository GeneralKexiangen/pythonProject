import random
import pymysql


def is_numeric(s):
    if s is None:
        return False
    try:
        float(s)
    except ValueError:
        return False
    else:
        return True


def sjss_and_df(b):
    # 设置指标得分的全局变量
    global rongyudefen, daikuandefen, anquandefen, tuntudefen, liantongdefen
    connect = pymysql.Connect(
        host='localhost',
        port=3306,
        user='root',
        password='root1234',
        db='test',
        charset='utf8'
    )

    row = 1

    from openpyxl import Workbook
    wb = Workbook()
    sheet1 = wb.active
    sheet1.cell(row=1, column=1, value='网络节点')
    sheet1.cell(row=1, column=2, value='A设备')
    sheet1.cell(row=1, column=3, value='B设备')
    sheet1.cell(row=1, column=4, value='C设备')
    sheet1.cell(row=1, column=5, value='D设备')
    sheet1.cell(row=1, column=6, value='E设备')
    sheet1.cell(row=1, column=7, value='F设备')
    sheet1.cell(row=1, column=8, value='G设备')
    sheet1.cell(row=1, column=9, value='H设备')
    sheet1.cell(row=1, column=10, value='I设备')
    sheet1.cell(row=1, column=11, value='J设备')
    sheet1.cell(row=1, column=12, value='节点等级')
    sheet1.cell(row=1, column=13, value='满足方案1？')
    sheet1.cell(row=1, column=14, value='满足方案2？')
    sheet1.cell(row=1, column=15, value='满足方案3？')
    sheet1.cell(row=1, column=16, value='满足方案4？')
    sheet1.cell(row=1, column=17, value='方案1指标平均得分')
    sheet1.cell(row=1, column=18, value='方案2指标平均得分')
    sheet1.cell(row=1, column=19, value='方案3指标平均得分')
    sheet1.cell(row=1, column=20, value='方案4指标平均得分')
    sheet1.cell(row=1, column=21, value='节点推荐方案顺序')

    # 读取数据库中部署方案数据信息
    cursor1 = connect.cursor()
    cursor2 = connect.cursor()
    cursor3 = connect.cursor()
    cursor1.execute("select * from 一级部署")
    cursor2.execute("select * from 二级部署")
    cursor3.execute("select * from 三级部署")
    list1 = []
    list2 = []
    list3 = []

    # 读取一级部署方案表
    for row1 in cursor1.fetchall():
        list1.append(row1)

    # 读取二级部署方案表
    for row1 in cursor2.fetchall():
        list2.append(row1)

    # 读取三级部署方案表
    for row1 in cursor3.fetchall():
        list3.append(row1)

    # 检索出一级部署方案中所有设备的数量
    yijifangan = [[0 for y in range(len(list1))] for x in range(len(list1[1]) - 4)]
    for j in range(len(list1[1]) - 4):
        for i in range(len(list1)):
            # yijifangan[j][i] = list1[i][j + 1]
            v = list1[i][j + 1]
            if not is_numeric(v):
                v = 0
            yijifangan[j][i] = int(v)

    # 检索出一级部署方案中的设备带宽、吞吐量、安全性
    yijidaikuan = []
    yijituntuliang = []
    yijianquanxing = []
    for i in range(len(list1)):
        yijidaikuan.append(int(list1[i][4]))
        yijituntuliang.append(int(list1[i][10]))
        yijianquanxing.append(int(list1[i][11]))

    # 检索出二级部署方案中所有设备的数量
    erjifangan = [[0 for y in range(len(list2))] for x in range(len(list2[1]) - 4)]
    for j in range(len(list2[1]) - 4):
        for i in range(len(list2)):
            # erjifangan[j][i] = list2[i][j + 1]
            v = list2[i][j + 1]
            if not is_numeric(v):
                v = 0
            erjifangan[j][i] = int(v)

    # 检索出二级部署方案中的设备带宽、吞吐量、安全性
    erjidaikuan = []
    erjituntuliang = []
    erjianquanxing = []
    for i in range(len(list2)):
        erjidaikuan.append(int(list2[i][4]))
        erjituntuliang.append(int(list2[i][10]))
        erjianquanxing.append(int(list2[i][11]))

    # 检索出三级部署方案中所有设备的数量
    sanjifangan = [[0 for y in range(len(list3))] for x in range(len(list3[1]) - 4)]
    for j in range(len(list3[1]) - 4):
        for i in range(len(list3)):
            # sanjifangan[j][i] = list3[i][j + 1]
            v = list3[i][j + 1]
            if not is_numeric(v):
                v = 0
            sanjifangan[j][i] = int(v)

    # 检索出三级部署方案中的设备带宽、吞吐量、安全性
    sanjidaikuan = []
    sanjituntuliang = []
    sanjianquanxing = []
    for i in range(len(list3)):
        sanjidaikuan.append(int(list3[i][4]))
        sanjituntuliang.append(int(list3[i][10]))
        sanjianquanxing.append(int(list3[i][11]))

    # 初始化
    jiedian = []
    dengji = ['一级', '二级', '三级']
    sl = [5, 6, 7, 8]
    shebei = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    # B = input("请输入网络中节点数量：B=")
    B = int(b)
    # 不同等级部署方案中，每个方案的设备数量
    shuliang = [[0 for y in range(10)] for x in range(int(B))]
    # 随机生成网络节点，格式为[节点名，等级，[各设备的数量]]
    jiedianxinxi = [[0 for y in range(3)] for x in range(int(B))]

    # 存储五个节点信息，格式为[[方案代号，方案得分],[]]
    A0 = []
    A1 = []
    A2 = []
    A3 = []
    A4 = []

    # 随机生成各网络节点等级、设备数量
    for i in range(int(B)):
        jiedian.append('A' + str(i))
        jiediangdengji = random.sample(dengji, 1)
        for j in range(10):
            b = random.sample(sl, 1)
            shuliang[i][j] = b[0]

        # 节点信息 节点+等级+设备数量
        jiedianxinxi[i][0] = jiedian[i]
        jiedianxinxi[i][1] = jiediangdengji[0]
        jiedianxinxi[i][2] = shuliang[i]

    # D = input("请输入网络中实际的带宽：D=")
    D = 3

    # T = input("请输入网络中实际的吞吐量：T=")
    T = 3

    # 进行遍历网络中所有节点并对各指标进行打分
    for i in range(len(jiedianxinxi)):

        # 满足的方案的位置
        mzfa = []
        # 满足方案的名称
        mzfa2 = []

        # 比较一级节点拥有的设备数量和一级部署方案需求的数量，如果符合需求则记录该方案
        if jiedianxinxi[i][1] == '一级':
            for j in range(len(yijifangan)):
                for x in range(10):
                    if jiedianxinxi[i][2][x] < yijifangan[j][x]:
                        break
                if x == 9:
                    mzfa.append(j)
        # 比较二级节点拥有的设备数量和二级部署方案需求的数量，如果符合需求则记录该方案
        if jiedianxinxi[i][1] == '二级':
            for j in range(len(erjifangan)):
                for x in range(10):
                    if jiedianxinxi[i][2][x] < erjifangan[j][x]:
                        break
                if x == 9:
                    mzfa.append(j)

        # 比较三级节点拥有的设备数量和三级部署方案需求的数量，如果符合需求则记录该方案
        if jiedianxinxi[i][1] == '三级':
            for j in range(len(sanjifangan)):
                for x in range(10):
                    if jiedianxinxi[i][2][x] < sanjifangan[j][x]:
                        break
                if x == 9:
                    mzfa.append(j)

        # 解析方案位置与方案名称
        for y in range(len(mzfa)):
            if mzfa[y] == 0:
                mzfa2.append("是")
            if mzfa[y] == 1:
                mzfa2.append("是")
            if mzfa[y] == 2:
                mzfa2.append("是")
            if mzfa[y] == 3:
                mzfa2.append("是")
            else:
                mzfa2.append("否")

        # 对设备冗余进行打分
        # 节点拥有的设备必须对满足对应等级的某一个部署方案
        if mzfa != []:

            # 一级方案冗余得分
            rongyudefen1 = []
            # 定义方案冗余得分的输出类型：[['1'，1.2],['2',1.1]]
            rongyudefen = [[0 for y in range(2)] for x in range(len(mzfa))]

            # 带宽打分
            daikuandefen1 = []
            daikuandefen = [[0 for y in range(2)] for x in range(len(mzfa))]

            # 吞吐量打分
            tuntudefen1 = []
            tuntudefen = [[0 for y in range(2)] for x in range(len(mzfa))]

            # 安全打分
            anquandefen1 = []
            anquandefen = [[0 for y in range(2)] for x in range(len(mzfa))]

            # 对节点连通性打分
            # 连通性参数liantongdefen,连通性有几个方案得几分，没有连通方案得0分
            liantongdefen = [[0 for y in range(2)] for x in range(len(mzfa))]
            for b in range(len(mzfa)):
                liantongdefen[b][0] = mzfa[b]
                liantongdefen[b][1] = 1

            # 一级打分
            if jiedianxinxi[i][1] == '一级':

                # 一级冗余打分
                for v in range(len(yijifangan)):
                    s = s1 = 0
                    shu0 = shu1 = 0
                    for m in range(10):
                        if yijifangan[v][m] != 0:
                            if yijifangan[v][m] <= jiedianxinxi[i][2][m] <= 3 * yijifangan[v][m]:
                                s1 = jiedianxinxi[i][2][m] * jiedianxinxi[i][2][m] / yijifangan[v][m]
                            if jiedianxinxi[i][2][m] > 3 * yijifangan[v][m]:
                                s1 = jiedianxinxi[i][2][m] * (
                                        3 - (jiedianxinxi[i][2][m] - yijifangan[v][m]) / 2 * yijifangan[v][m])
                            s = s + s1
                            shu0 = jiedianxinxi[i][2][m]
                            shu1 = shu1 + shu0
                    res1 = 0
                    if shu1 != 0 and shu1 is not None:
                        res1 = s / shu1
                    rongyudefen1.append(res1)
                    # rongyudefen1.append(s / shu1)
                for b in range(len(mzfa)):
                    rongyudefen[b][0] = mzfa[b]
                    rongyudefen[b][1] = rongyudefen1[mzfa[b]]

                # 一级带宽得分
                for j in range(len(yijifangan)):
                    s = s1 = 0
                    shu0 = shu1 = 0
                    for c in range(10):
                        if yijifangan[j][c] != 0 and yijidaikuan[c] <= int(D):
                            if yijidaikuan[c] < int(D):
                                s1 = 2 * yijifangan[j][c] * (int(D) / yijidaikuan[c])
                            if yijidaikuan[c] == int(D):
                                s1 = 1 * yijifangan[j][c]
                            s = s + s1
                            shu0 = yijifangan[j][c]
                            shu1 = shu1 + shu0
                    res1 = 0
                    if shu1 != 0 and shu1 is not None:
                        res1 = s / shu1
                    daikuandefen1.append(res1)
                    # daikuandefen1.append(s / shu1)
                for b in range(len(mzfa)):
                    daikuandefen[b][0] = mzfa[b]
                    daikuandefen[b][1] = daikuandefen1[mzfa[b]]

                # 一级吞吐量打分
                for j in range(len(yijifangan)):
                    s = s1 = 0
                    shu0 = shu1 = 0
                    for c in range(10):
                        if yijifangan[j][c] != 0:
                            if yijituntuliang[c] > int(T):
                                s1 = 0.5 * yijifangan[j][c]
                            if yijituntuliang[c] < int(T):
                                s1 = 1.5 * yijifangan[j][c] * (int(T) / yijituntuliang[c])
                            if yijituntuliang[c] == int(T):
                                s1 = 1 * yijifangan[j][c]
                            s = s + s1
                            shu0 = yijifangan[j][c]
                            shu1 = shu1 + shu0
                    res1 = 0
                    if shu1 != 0 and shu1 is not None:
                        res1 = s / shu1
                    tuntudefen1.append(res1)
                    # tuntudefen1.append(s / shu1)
                for b in range(len(mzfa)):
                    tuntudefen[b][0] = mzfa[b]
                    tuntudefen[b][1] = tuntudefen1[mzfa[b]]

                # 一级安全性打分
                for j in range(len(yijifangan)):
                    s = s1 = 0
                    shu0 = shu1 = 0
                    for c in range(10):
                        if yijifangan[j][c] != 0:
                            if yijianquanxing[c] > 1:
                                s1 = -0.5 * yijifangan[j][c]
                            if yijianquanxing[c] == 1:
                                s1 = 1 * yijifangan[j][c]
                            if yijianquanxing[c] < 1:
                                s1 = 2 * yijifangan[j][c]
                            s = s + s1
                            shu0 = yijifangan[j][c]
                            shu1 = shu1 + shu0
                    res1 = 0
                    if shu1 != 0 and shu1 is not None:
                        res1 = s / shu1
                    anquandefen1.append(res1)
                    # anquandefen1.append(s / shu1)
                for b in range(len(mzfa)):
                    anquandefen[b][0] = mzfa[b]
                    anquandefen[b][1] = anquandefen1[mzfa[b]]

            # 二级方案打分
            if jiedianxinxi[i][1] == '二级':
                # 二级冗余打分
                for v in range(len(erjifangan)):
                    s = s1 = 0
                    shu0 = shu1 = 0
                    for m in range(10):
                        if erjifangan[v][m] != 0:
                            if erjifangan[v][m] <= jiedianxinxi[i][2][m] <= 3 * erjifangan[v][m]:
                                s1 = jiedianxinxi[i][2][m] * jiedianxinxi[i][2][m] / erjifangan[v][m]
                            if jiedianxinxi[i][2][m] > 3 * erjifangan[v][m]:
                                s1 = jiedianxinxi[i][2][m] * (3 - (jiedianxinxi[i][2][m] - erjifangan[v][m]) / 2)
                            s = s + s1
                            shu0 = jiedianxinxi[i][2][m]
                            shu1 = shu1 + shu0
                    res1 = 0
                    if shu1 != 0 and shu1 is not None:
                        res1 = s / shu1
                    rongyudefen1.append(res1)
                    # rongyudefen1.append(s / shu1)
                for b in range(len(mzfa)):
                    rongyudefen[b][0] = mzfa[b]
                    rongyudefen[b][1] = rongyudefen1[mzfa[b]]

                # 二级带宽打分
                for j in range(len(erjifangan)):
                    s = s1 = 0
                    shu0 = shu1 = 0
                    for c in range(10):
                        if erjifangan[j][c] != 0 and erjidaikuan[c] <= int(D):
                            if erjidaikuan[c] < int(D):
                                s1 = 2 * erjifangan[j][c] * (int(D) / erjidaikuan[c])
                            if erjidaikuan[c] == int(D):
                                s1 = 1 * erjifangan[j][c]
                            s = s + s1
                            shu0 = erjifangan[j][c]
                            shu1 = shu1 + shu0
                    res1 = 0
                    if shu1 != 0 and shu1 is not None:
                        res1 = s / shu1
                    daikuandefen1.append(res1)
                    # daikuandefen1.append(s / shu1)
                for b in range(len(mzfa)):
                    daikuandefen[b][0] = mzfa[b]
                    daikuandefen[b][1] = daikuandefen1[mzfa[b]]

                # 二级吞吐量打分
                for j in range(len(erjifangan)):
                    s = s1 = 0
                    shu0 = shu1 = 0
                    for c in range(10):
                        if erjifangan[j][c] != 0:
                            if erjituntuliang[c] > int(T):
                                s1 = 0.5 * erjifangan[j][c]
                            if erjituntuliang[c] < int(T):
                                s1 = 1.5 * erjifangan[j][c] * (int(T) / erjituntuliang[c])
                            if erjituntuliang[c] == int(T):
                                s1 = 1 * erjifangan[j][c]
                            s = s + s1
                            shu0 = erjifangan[j][c]
                            shu1 = shu1 + shu0
                    res1 = 0
                    if shu1 != 0 and shu1 is not None:
                        res1 = s / shu1
                    tuntudefen1.append(res1)
                    # tuntudefen1.append(s / shu1)
                for b in range(len(mzfa)):
                    tuntudefen[b][0] = mzfa[b]
                    tuntudefen[b][1] = tuntudefen1[mzfa[b]]

                # 二级安全性打分
                for j in range(len(erjifangan)):
                    s = s1 = 0
                    shu0 = shu1 = 0
                    for c in range(10):
                        if erjifangan[j][c] != 0:
                            if erjianquanxing[c] > 1:
                                s1 = -0.5 * erjifangan[j][c]
                            if erjianquanxing[c] == 1:
                                s1 = 1 * erjifangan[j][c]
                            if erjianquanxing[c] < 1:
                                s1 = 2 * erjifangan[j][c]
                            s = s + s1
                            shu0 = erjifangan[j][c]
                            shu1 = shu1 + shu0
                    res1 = 0
                    if shu1 != 0 and shu1 is not None:
                        res1 = s / shu1
                    anquandefen1.append(res1)
                    # anquandefen1.append(s / shu1)
                for b in range(len(mzfa)):
                    anquandefen[b][0] = mzfa[b]
                    anquandefen[b][1] = anquandefen1[mzfa[b]]

            # 三级方案打分
            if jiedianxinxi[i][1] == '三级':
                # 三级冗余打分
                for v in range(len(sanjifangan)):
                    s = s1 = 0
                    shu0 = shu1 = 0
                    for m in range(10):
                        if sanjifangan[v][m] != 0:
                            if sanjifangan[v][m] <= jiedianxinxi[i][2][m] <= 3 * sanjifangan[v][m]:
                                s1 = jiedianxinxi[i][2][m] * jiedianxinxi[i][2][m] / sanjifangan[v][m]
                            if jiedianxinxi[i][2][m] > 3 * sanjifangan[v][m]:
                                s1 = jiedianxinxi[i][2][m] * (3 - (jiedianxinxi[i][2][m] - sanjifangan[v][m]) / 2)
                            s = s + s1
                            shu0 = jiedianxinxi[i][2][m]
                            shu1 = shu1 + shu0
                    res1 = 0
                    if shu1 != 0 and shu1 is not None:
                        res1 = s / shu1
                    rongyudefen1.append(res1)
                    # rongyudefen1.append(s / shu1)
                for b in range(len(mzfa)):
                    rongyudefen[b][0] = mzfa[b]
                    rongyudefen[b][1] = rongyudefen1[mzfa[b]]

                # 三级带宽打分
                for j in range(len(sanjifangan)):
                    s = s1 = 0
                    shu0 = shu1 = 0
                    for c in range(10):
                        if sanjifangan[j][c] != 0 and sanjidaikuan[c] <= int(D):
                            if sanjidaikuan[c] < int(D):
                                s1 = 2 * sanjifangan[j][c] * (int(D) / sanjidaikuan[c])
                            if sanjidaikuan[c] == int(D):
                                s1 = 1 * sanjifangan[j][c]
                            s = s + s1
                            shu0 = sanjifangan[j][c]
                            shu1 = shu1 + shu0
                    res1 = 0
                    if shu1 != 0 and shu1 is not None:
                        res1 = s / shu1
                    daikuandefen1.append(res1)
                    # daikuandefen1.append(s / shu1)
                for b in range(len(mzfa)):
                    daikuandefen[b][0] = mzfa[b]
                    daikuandefen[b][1] = daikuandefen1[mzfa[b]]

                # 三级吞吐量打分
                for j in range(len(sanjifangan)):
                    s = s1 = 0
                    shu0 = shu1 = 0
                    for c in range(10):
                        if sanjifangan[j][c] != 0:
                            if sanjituntuliang[c] > int(T):
                                s1 = 0.5 * sanjifangan[j][c]
                            if sanjituntuliang[c] < int(T):
                                s1 = 1.5 * sanjifangan[j][c] * (int(T) / sanjituntuliang[c])
                            if sanjituntuliang[c] == int(T):
                                s1 = 1 * sanjifangan[j][c]
                            s = s + s1
                            shu0 = sanjifangan[j][c]
                            shu1 = shu1 + shu0
                    res1 = 0
                    if shu1 != 0 and shu1 is not None:
                        res1 = s / shu1
                    tuntudefen1.append(res1)
                    # tuntudefen1.append(s / shu1)
                for b in range(len(mzfa)):
                    tuntudefen[b][0] = mzfa[b]
                    tuntudefen[b][1] = tuntudefen1[mzfa[b]]

                # 三级安全性打分
                for j in range(len(sanjifangan)):
                    s = s1 = 0
                    shu0 = shu1 = 0
                    for c in range(10):
                        if sanjifangan[j][c] != 0:
                            if sanjianquanxing[c] > 1:
                                s1 = -0.5 * sanjifangan[j][c]
                            if sanjianquanxing[c] == 1:
                                s1 = 1 * sanjifangan[j][c]
                            if sanjianquanxing[c] < 1:
                                s1 = 2 * sanjifangan[j][c]
                            s = s + s1
                            shu0 = sanjifangan[j][c]
                            shu1 = shu1 + shu0
                    res1 = 0
                    if shu1 != 0 and shu1 is not None:
                        res1 = s / shu1
                    anquandefen1.append(res1)
                    # anquandefen1.append(s / shu1)
                for b in range(len(mzfa)):
                    anquandefen[b][0] = mzfa[b]
                    anquandefen[b][1] = anquandefen1[mzfa[b]]

        # 总分
        zongfeng = [[0 for y in range(2)] for x in range(len(mzfa))]
        for e in range(len(mzfa)):
            zongfeng[e][0] = mzfa[e]
            zongfeng[e][1] = rongyudefen[e][1] + daikuandefen[e][1] + anquandefen[e][1] + tuntudefen[e][1] + \
                             liantongdefen[e][1]

        # 对同一等级部署方案中的各方案进行得分排序，得分高的优先推荐
        zongfengpaixu = zongfeng
        for k in range(len(zongfengpaixu)):
            for h in range(k + 1, len(zongfengpaixu)):
                if zongfengpaixu[k][1] < zongfengpaixu[h][1]:
                    a = zongfengpaixu[k]
                    zongfengpaixu[k] = zongfengpaixu[h]
                    zongfengpaixu[h] = a

        # 将zongfengpaixu进行赋值
        if i == 0:
            A0 = zongfengpaixu
        if i == 1:
            A1 = zongfengpaixu
        if i == 2:
            A2 = zongfengpaixu
        if i == 3:
            A3 = zongfengpaixu
        if i == 4:
            A4 = zongfengpaixu

        # 输出节点得分排序后的节点方案顺序
        tr = []
        for c in range(len(zongfengpaixu)):
            if zongfengpaixu[c][0] == 0:
                tr.append("方案1")
            if zongfengpaixu[c][0] == 1:
                tr.append("方案2")
            if zongfengpaixu[c][0] == 2:
                tr.append("方案3")
            if zongfengpaixu[c][0] == 3:
                tr.append("方案4")

        # arr 用于存储节点方案得分、推荐顺序
        arr = [[0 for y in range(19)] for x in range(len(jiedianxinxi))]
        for c in range(len(jiedianxinxi)):
            arr[c][0] = jiedianxinxi[i][0]
            arr[c][1] = jiedianxinxi[i][2][0]
            arr[c][2] = jiedianxinxi[i][2][1]
            arr[c][3] = jiedianxinxi[i][2][2]
            arr[c][4] = jiedianxinxi[i][2][3]
            arr[c][5] = jiedianxinxi[i][2][4]
            arr[c][6] = jiedianxinxi[i][2][5]
            arr[c][7] = jiedianxinxi[i][2][6]
            arr[c][8] = jiedianxinxi[i][2][7]
            arr[c][9] = jiedianxinxi[i][2][8]
            arr[c][10] = jiedianxinxi[i][2][9]

            # 判断方案是否满足
            for f in range(len(mzfa)):
                if 0 in mzfa:
                    arr[c][11] = '是'
                if 1 in mzfa:
                    arr[c][12] = '是'
                if 2 in mzfa:
                    arr[c][13] = '是'
                if 3 in mzfa:
                    arr[c][14] = '是'

                # 将对应节点的推荐方案赋值
                if zongfeng[f][0] == 0:
                    arr[c][15] = zongfeng[f][1] / 5
                if zongfeng[f][0] == 1:
                    arr[c][16] = zongfeng[f][1] / 5
                if zongfeng[f][0] == 2:
                    arr[c][17] = zongfeng[f][1] / 5
                if zongfeng[f][0] == 3:
                    arr[c][18] = zongfeng[f][1] / 5

        # 输出
        sheet1.cell(row=row + 1, column=1, value=jiedianxinxi[i][0])
        sheet1.cell(row=row + 1, column=2, value=jiedianxinxi[i][2][0])
        sheet1.cell(row=row + 1, column=3, value=jiedianxinxi[i][2][1])
        sheet1.cell(row=row + 1, column=4, value=jiedianxinxi[i][2][2])
        sheet1.cell(row=row + 1, column=5, value=jiedianxinxi[i][2][3])
        sheet1.cell(row=row + 1, column=6, value=jiedianxinxi[i][2][4])
        sheet1.cell(row=row + 1, column=7, value=jiedianxinxi[i][2][5])
        sheet1.cell(row=row + 1, column=8, value=jiedianxinxi[i][2][6])
        sheet1.cell(row=row + 1, column=9, value=jiedianxinxi[i][2][7])
        sheet1.cell(row=row + 1, column=10, value=jiedianxinxi[i][2][8])
        sheet1.cell(row=row + 1, column=11, value=jiedianxinxi[i][2][9])
        sheet1.cell(row=row + 1, column=12, value=jiedianxinxi[i][1])
        sheet1.cell(row=row + 1, column=13, value=arr[c][11])
        sheet1.cell(row=row + 1, column=14, value=arr[c][12])
        sheet1.cell(row=row + 1, column=15, value=arr[c][13])
        sheet1.cell(row=row + 1, column=16, value=arr[c][14])
        sheet1.cell(row=row + 1, column=17, value="%.2f" % arr[c][15])
        sheet1.cell(row=row + 1, column=18, value="%.2f" % arr[c][16])
        sheet1.cell(row=row + 1, column=19, value="%.2f" % arr[c][17])
        sheet1.cell(row=row + 1, column=20, value="%.2f" % arr[c][18])
        sheet1.cell(row=row + 1, column=21, value=f"{tr}")
        row += 1
        wb.save('网络各节点检索结果.xlsx')

    # 进行整个网络的节点方案连接，通过在每个节点中随机的找出一个满足部署方案与其他节点进行组合，并给出整体得分
    # 进行随机组合
    index_set = []
    temp_set = []
    result_set1 = []
    for q in A0:
        for p in A1:
            for r in A2:
                for s in A3:
                    for t in A4:
                        index_set.append(q)
                        index_set.append(p)
                        index_set.append(r)
                        index_set.append(s)
                        index_set.append(t)
    for c in range(0, len(index_set), 5):
        temp_set.append(index_set[c])
        temp_set.append(index_set[c + 1])
        temp_set.append(index_set[c + 2])
        temp_set.append(index_set[c + 3])
        temp_set.append(index_set[c + 4])
        result_set1.append(temp_set)
        temp_set = []
    # print(result_set1)

    # 对n个网络整体方案进行得分排序，得分高的优化推荐
    result_set = result_set1
    zhengtifangandefen1 = []

    # 将每个方案的得分进行累加
    for c in range(len(result_set)):
        zhengtifangandefen1.append(
            result_set[c][0][1] + result_set[c][1][1] + result_set[c][2][1] + result_set[c][3][1] + result_set[c][4][1])

    # 对方案进行排序
    for k in range(len(result_set)):
        for h in range(k + 1, len(result_set)):
            if zhengtifangandefen1[k] < zhengtifangandefen1[h]:
                a1 = result_set[k]
                result_set[k] = result_set[h]
                result_set[h] = a1

    # 对方案得分进行排序
    zhengtifangandefen1.sort(reverse=True)

    # 判断各节点部署方案能够连接成整体方案
    if result_set1 != []:
        from openpyxl import Workbook
        wb = Workbook()
        sheet1 = wb.active
        sheet1.cell(row=1, column=1, value='整体方案编号')
        sheet1.cell(row=1, column=2, value='A')
        sheet1.cell(row=1, column=3, value='<-->')
        sheet1.cell(row=1, column=4, value='B')
        sheet1.cell(row=1, column=6, value='A')
        sheet1.cell(row=1, column=7, value='<-->')
        sheet1.cell(row=1, column=8, value='C')
        sheet1.cell(row=1, column=10, value='B')
        sheet1.cell(row=1, column=11, value='<-->')
        sheet1.cell(row=1, column=12, value='C')
        sheet1.cell(row=1, column=14, value='C')
        sheet1.cell(row=1, column=15, value='<-->')
        sheet1.cell(row=1, column=16, value='D')
        sheet1.cell(row=1, column=18, value='C')
        sheet1.cell(row=1, column=19, value='<-->')
        sheet1.cell(row=1, column=20, value='E')
        sheet1.cell(row=1, column=21, value="网络节点整体部署方案节点平均得分")

        # row进行初始化
        row = row - 5
        print("已经成功的将所有节点部署方案进行随机进行组合，并按照网络整体部署方案得分从高到底存储的表格中！！！")
        for c in range(len(result_set)):
            ter = []
            for i in range(5):
                if result_set[c][i][0] == 0:
                    ter.append('方案1')
                if result_set[c][i][0] == 1:
                    ter.append('方案2')
                if result_set[c][i][0] == 2:
                    ter.append('方案3')
                if result_set[c][i][0] == 3:
                    ter.append('方案4')
            while len(ter) < 5:
                ter.append('方案4')

            # 输出整体部署方案
            sheet1.cell(row=row + 1, column=1, value=c + 1)
            sheet1.cell(row=row + 1, column=2, value=ter[0])
            sheet1.cell(row=row + 1, column=4, value=ter[1])
            sheet1.cell(row=row + 1, column=6, value=ter[0])
            sheet1.cell(row=row + 1, column=8, value=ter[2])
            sheet1.cell(row=row + 1, column=10, value=ter[1])
            sheet1.cell(row=row + 1, column=12, value=ter[2])
            sheet1.cell(row=row + 1, column=14, value=ter[2])
            sheet1.cell(row=row + 1, column=16, value=ter[3])
            sheet1.cell(row=row + 1, column=18, value=ter[2])
            sheet1.cell(row=row + 1, column=20, value=ter[4])
            sheet1.cell(row=row + 1, column=21, value="%.2f" % (zhengtifangandefen1[c] / 5))
            row += 1
            wb.save('网络应用系统的部署方案及方案总分.xlsx')
    else:
        print("网络中某个节点的设备无法满足对应等级的方案要求，无法形成网络整体方案，请检查节点设备数量情况！！！")


if __name__ == '__main__':
    sjss_and_df('5')
